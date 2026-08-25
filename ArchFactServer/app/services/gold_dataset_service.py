from __future__ import annotations

import hashlib
import re
import unicodedata
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from app.core.config import Settings
from app.core.errors import ConflictError, DomainError
from app.repositories.mongo_repository import MongoRepository

FIELD_MAPPING = {
    "尺寸": "measurements",
    "外形描述": "morphological_description",
    "颜色": "surface_color",
    "器型1": "artifact_group",
    "器型2": "category",
    "型别": "type",
    "式别": "subtype",
    "材质": "texture",
    "完整程度": "completeness",
    "图注": "figure_caption",
    "地层": "stratigraphy",
    "备注": "notes",
}

REGION_KIND = {
    0: "artifact",
    1: "number",
    2: "caption",
    3: "grave_drawing",
    4: "group",
}

IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".webp", ".bmp"}
EMPTY_MARKERS = {"", "-", "—", "－", "无", "none", "null", "nan"}


def utc_now() -> datetime:
    return datetime.now(UTC)


def stable_id(prefix: str, *parts: object) -> str:
    raw = "\x1f".join(str(part) for part in parts)
    return f"{prefix}_{hashlib.sha1(raw.encode('utf-8')).hexdigest()[:24]}"


def clean_value(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = unicodedata.normalize("NFKC", str(value)).strip()
    return None if text.lower() in EMPTY_MARKERS else text


def normalize_identifier(value: Any) -> str:
    text = clean_value(value) or ""
    text = text.replace("，", ",").replace("：", ",").replace(":", ",")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r",+", ",", text).strip(",")
    return text.upper()


def canonical_artifact_id(context: Any, sequence: Any) -> str:
    parts = (clean_value(context), clean_value(sequence))
    return normalize_identifier(",".join(part for part in parts if part))


def normalize_reference(value: str) -> str:
    text = unicodedata.normalize("NFKC", value)
    text = text.replace("，", ",").replace("：", ",").replace(":", ",")
    text = text.replace("－", "-").replace("—", "-")
    return re.sub(r"\s+", "", text).strip(".,;；")


def extract_color_plate_keys(value: Any) -> list[str]:
    text = clean_value(value) or ""
    matches = re.findall(
        r"彩版[〇零一二三四五六七八九十百0-9O○\-]+(?:[,，][〇零一二三四五六七八九十百0-9]+)?",
        text,
        flags=re.IGNORECASE,
    )
    return list(dict.fromkeys(normalize_reference(match) for match in matches))


class GoldDatasetService:
    """Imports human annotations into an isolated evaluation namespace.

    These documents are never queried by the extraction pipeline. They are only
    resolved after a human verification session has been submitted.
    """

    def __init__(self, settings: Settings, repository: MongoRepository) -> None:
        self._settings = settings
        self._repository = repository

    async def import_wenjiashan(
        self,
        *,
        document_id: str,
        version: str,
        replace: bool,
    ) -> dict[str, Any]:
        document = await self._repository.get_document(document_id)
        root = self._resolve_root()
        source_document_verified = self._verify_source_document(root, document)
        workbook_path = self._find_workbook(root)
        labels_root = root / "labels"
        assets_root = root / "cutpictures"
        if not labels_root.is_dir() or not assets_root.is_dir():
            raise DomainError("人工标注数据目录缺少 labels 或 cutpictures 子目录")

        existing = await self._repository.get_gold_dataset_for_document(
            document_id=document_id,
            version=version,
        )
        if existing and not replace:
            raise ConflictError("该 PDF 的同版本人工标注数据已导入；如需重建请使用 replace=true")

        dataset_id = existing["_id"] if existing else stable_id("gold", document_id, version)
        records = self._read_records(workbook_path, dataset_id)
        assets = self._read_assets(assets_root, root, dataset_id, records)
        regions = self._read_regions(labels_root, dataset_id)
        links = self._build_asset_links(dataset_id, records, assets)

        artifact_matches = len(
            {link["record_id"] for link in links if link["link_type"] == "artifact_crop"}
        )
        plate_matches = len(
            {link["record_id"] for link in links if link["link_type"] == "color_plate"}
        )
        warnings: list[str] = []
        if source_document_verified is None:
            warnings.append("未找到配套源 PDF，人工标注数据绑定未执行文件指纹确认")
        artifact_expected = sum(bool(record["canonical_artifact_id"]) for record in records)
        if artifact_matches < artifact_expected:
            warnings.append(f"{artifact_expected - artifact_matches} 条记录未找到同名器物裁剪图")
        records_with_plate = sum(bool(record.get("color_plate_keys")) for record in records)
        if plate_matches < records_with_plate:
            warnings.append(f"{records_with_plate - plate_matches} 条彩版引用未找到直接文件匹配")

        now = utc_now()
        dataset = {
            "_id": dataset_id,
            "name": "文家山人工标注数据",
            "document_id": document_id,
            "document_filename": document.get("filename", ""),
            "version": version,
            "status": "ready",
            "source_type": "human_annotation",
            "source_root_key": "wenjiashan_gold_dataset",
            "source_document_verified": source_document_verified is True,
            "workbook_name": workbook_path.name,
            "record_count": len(records),
            "region_count": len(regions),
            "asset_count": len(assets),
            "link_count": len(links),
            "matched_artifact_assets": artifact_matches,
            "matched_color_plate_assets": plate_matches,
            "warnings": warnings,
            "created_at": existing.get("created_at", now) if existing else now,
            "updated_at": now,
        }
        await self._repository.replace_gold_dataset(
            dataset=dataset,
            records=records,
            regions=regions,
            assets=assets,
            links=links,
        )
        return dataset

    def _resolve_root(self) -> Path:
        configured = self._settings.gold_dataset_root
        if configured is None:
            raise DomainError("尚未配置 GOLD_DATASET_ROOT")
        configured = configured.expanduser()
        project_root = Path(__file__).resolve().parents[2]
        root = (
            configured.resolve()
            if configured.is_absolute()
            else (project_root / configured).resolve()
        )
        if not root.is_dir():
            raise DomainError(f"人工标注数据目录不存在：{root}")
        return root

    @staticmethod
    def _find_workbook(root: Path) -> Path:
        candidates = sorted(path for path in root.glob("*.xlsx") if not path.name.startswith("~$"))
        if not candidates:
            raise DomainError("人工标注数据目录中没有找到 xlsx 文件")
        return candidates[0]

    @staticmethod
    def _verify_source_document(root: Path, document: dict[str, Any]) -> bool | None:
        candidates = [
            path
            for path in root.parent.glob("*.pdf")
            if "文家山" in path.name and path.is_file()
        ]
        if not candidates:
            return None
        source_pdf = sorted(candidates, key=lambda path: path.name)[0]
        digest = hashlib.sha256()
        with source_pdf.open("rb") as stream:
            for chunk in iter(lambda: stream.read(1024 * 1024), b""):
                digest.update(chunk)
        if digest.hexdigest() != document.get("sha256"):
            raise ConflictError("上传的 PDF 与文家山人工标注数据配套源文件指纹不一致，已拒绝绑定")
        return True

    @staticmethod
    def _read_records(workbook_path: Path, dataset_id: str) -> list[dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - deployment dependency guard
            raise DomainError("缺少 openpyxl，无法导入人工标注 Excel") from exc

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        headers = [clean_value(value) or "" for value in next(rows)]
        records: list[dict[str, Any]] = []
        for source_row, values in enumerate(rows, start=2):
            raw = dict(zip(headers, values, strict=False))
            if not any(clean_value(value) for value in values):
                continue
            context = clean_value(raw.get("遗迹号"))
            sequence = clean_value(raw.get("序号"))
            artifact_id = canonical_artifact_id(context, sequence)
            fields = {
                target: clean_value(raw.get(source))
                for source, target in FIELD_MAPPING.items()
            }
            fields["artifact_id"] = artifact_id or None
            records.append(
                {
                    "_id": stable_id("goldrec", dataset_id, source_row),
                    "dataset_id": dataset_id,
                    "source_row": source_row,
                    "context_id": context,
                    "sequence_no": sequence,
                    "canonical_artifact_id": artifact_id,
                    "fields": fields,
                    "color_plate_keys": extract_color_plate_keys(raw.get("图注")),
                    "raw": {key: clean_value(value) for key, value in raw.items()},
                }
            )
        workbook.close()
        return records

    @staticmethod
    def _read_assets(
        assets_root: Path,
        root: Path,
        dataset_id: str,
        records: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        record_ids = {record["canonical_artifact_id"] for record in records}
        assets: list[dict[str, Any]] = []
        for path in sorted(assets_root.rglob("*")):
            if not path.is_file() or path.suffix.lower() not in IMAGE_SUFFIXES:
                continue
            normalized_stem = normalize_identifier(path.stem)
            reference_keys = [normalize_reference(path.stem)]
            if path.stem.startswith("彩版"):
                asset_type = "color_plate"
            elif normalized_stem in record_ids:
                asset_type = "artifact_crop"
                reference_keys.append(normalized_stem)
            else:
                asset_type = "reference_image"
            relative_path = path.relative_to(root).as_posix()
            assets.append(
                {
                    "_id": stable_id("goldasset", dataset_id, relative_path),
                    "dataset_id": dataset_id,
                    "asset_type": asset_type,
                    "object_key": relative_path,
                    "filename": path.name,
                    "normalized_stem": normalized_stem,
                    "reference_keys": list(dict.fromkeys(reference_keys)),
                    "size": path.stat().st_size,
                }
            )
        return assets

    @staticmethod
    def _read_regions(labels_root: Path, dataset_id: str) -> list[dict[str, Any]]:
        regions: list[dict[str, Any]] = []
        for path in sorted(labels_root.glob("*.txt")):
            page_match = re.search(r"(\d+)$", path.stem)
            if not page_match:
                continue
            page = int(page_match.group(1))
            lines = path.read_text(encoding="utf-8-sig").splitlines()
            for position, line in enumerate(lines, start=1):
                parts = line.strip().split()
                if len(parts) < 5:
                    continue
                try:
                    class_id = int(parts[0])
                    x_center, y_center, width, height = map(float, parts[1:5])
                except ValueError:
                    continue
                kind = REGION_KIND.get(class_id, "other")
                bbox = [
                    max(0.0, x_center - width / 2),
                    max(0.0, y_center - height / 2),
                    min(1.0, x_center + width / 2),
                    min(1.0, y_center + height / 2),
                ]
                regions.append(
                    {
                        "_id": stable_id("goldregion", dataset_id, page, position),
                        "dataset_id": dataset_id,
                        "page": page,
                        "kind": kind,
                        "class_id": class_id,
                        "bbox": bbox,
                        "source_file": path.name,
                        "source_line": position,
                    }
                )
        return regions

    @staticmethod
    def _build_asset_links(
        dataset_id: str,
        records: list[dict[str, Any]],
        assets: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        artifact_assets = {
            asset["normalized_stem"]: asset
            for asset in assets
            if asset["asset_type"] == "artifact_crop"
        }
        plate_assets: dict[str, list[dict[str, Any]]] = {}
        for asset in assets:
            if asset["asset_type"] != "color_plate":
                continue
            for key in asset["reference_keys"]:
                plate_assets.setdefault(key, []).append(asset)

        links: list[dict[str, Any]] = []
        for record in records:
            artifact_asset = artifact_assets.get(record["canonical_artifact_id"])
            if artifact_asset:
                links.append(
                    {
                        "_id": stable_id("goldlink", record["_id"], artifact_asset["_id"]),
                        "dataset_id": dataset_id,
                        "record_id": record["_id"],
                        "asset_id": artifact_asset["_id"],
                        "link_type": "artifact_crop",
                        "status": "confirmed",
                        "method": "normalized_filename_exact",
                        "confidence": 1.0,
                    }
                )
            for plate_key in record.get("color_plate_keys", []):
                for plate_asset in plate_assets.get(plate_key, []):
                    links.append(
                        {
                            "_id": stable_id("goldlink", record["_id"], plate_asset["_id"]),
                            "dataset_id": dataset_id,
                            "record_id": record["_id"],
                            "asset_id": plate_asset["_id"],
                            "link_type": "color_plate",
                            "status": "confirmed",
                            "method": "figure_reference_exact",
                            "confidence": 1.0,
                        }
                    )
        return links
